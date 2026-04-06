from datetime import datetime
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.models.machine_schedule_result import MachineScheduleResult
from app.models.order_schedule_snapshot import OrderScheduleSnapshot
from app.models.part_schedule_result import PartScheduleResult
from app.services.schedule_export_service import ExportService


@pytest.mark.asyncio
async def test_export_machine_schedules(app_client, db_session):
    db_session.add_all(
        [
            MachineScheduleResult(
                order_line_id=1,
                contract_no="HT001",
                product_model="MC1-80",
                schedule_status="scheduled_stale",
                drawing_released=True,
                warning_level="abnormal",
                machine_cycle_days=Decimal("60"),
                machine_assembly_days=Decimal("3"),
            ),
            MachineScheduleResult(
                order_line_id=2,
                contract_no="HT999",
                product_model="MC9-90",
                schedule_status="scheduled",
                drawing_released=False,
                warning_level="normal",
                machine_cycle_days=Decimal("30"),
                machine_assembly_days=Decimal("2"),
            ),
        ]
    )
    await db_session.commit()

    resp = await app_client.get("/api/exports/machine-schedules?sort_field=contract_no&sort_order=desc")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert len(resp.content) > 0
    assert "filename*=UTF-8''" in resp.headers["content-disposition"]
    assert "%E6%95%B4%E6%9C%BA%E6%8E%92%E4%BA%A7%E5%88%97%E8%A1%A8_" in resp.headers["content-disposition"]

    workbook = load_workbook(BytesIO(resp.content))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    first_row = [cell.value for cell in sheet[2]]
    second_row = [cell.value for cell in sheet[3]]

    assert headers[-1] == "异常标识"
    assert "整机主周期(天)" in headers
    assert "整机总装时长(天)" in headers
    assert "排产状态" in headers
    assert "发图状态" in headers
    assert first_row[1] == "HT999"
    assert second_row[1] == "HT001"
    assert "待重排" in second_row
    assert "已发图" in second_row
    assert "异常" in second_row


@pytest.mark.asyncio
async def test_export_machine_schedules_includes_extended_sales_fields(app_client, db_session):
    db_session.add(
        OrderScheduleSnapshot(
            order_line_id=8,
            contract_no="HT008",
            customer_name="客户8",
            product_series="MC",
            product_model="MC-008",
            product_name="高性能冲床",
            material_no="MAT-008",
            plant="1100",
            quantity=Decimal("1"),
            order_type="3",
            line_total_amount=Decimal("123456"),
            order_date=datetime(2026, 3, 8, 0, 0, 0),
            business_group="精机",
            custom_no="DZ-008",
            sales_person_name="李四",
            sales_branch_company="北方分公司",
            sales_sub_branch="烟台支公司",
            order_no="SO008",
            sap_code="SAP008",
            sap_line_no="000020",
            confirmed_delivery_date=datetime(2026, 4, 8, 0, 0, 0),
            drawing_released=True,
            drawing_release_date=datetime(2026, 3, 10, 0, 0, 0),
            custom_requirement="非标台面",
            review_comment="允许排产",
            trigger_date=datetime(2026, 3, 20, 0, 0, 0),
            schedule_status="scheduled",
            planned_start_date=datetime(2026, 3, 22, 0, 0, 0),
            planned_end_date=datetime(2026, 4, 8, 0, 0, 0),
            machine_cycle_days=Decimal("18"),
            machine_assembly_days=Decimal("3"),
            warning_level="normal",
        )
    )
    await db_session.commit()

    resp = await app_client.get("/api/exports/machine-schedules")
    workbook = load_workbook(BytesIO(resp.content))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    row = [cell.value for cell in sheet[2]]
    values = dict(zip(headers, row, strict=False))

    assert "合同金额" in headers
    assert "事业群" in headers
    assert "定制要求" in headers
    assert values["订单类型"] == "定制"
    assert values["合同金额"] == 123456
    assert values["事业群"] == "精机"
    assert values["分公司"] == "北方分公司"
    assert values["支公司"] == "烟台支公司"
    assert values["定制要求"] == "非标台面"
    assert values["评审意见"] == "允许排产"


@pytest.mark.asyncio
async def test_export_part_schedules_empty(app_client):
    resp = await app_client.get("/api/exports/part-schedules")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]


@pytest.mark.asyncio
async def test_export_part_schedules_supports_order_line_id_filter(app_client, db_session):
    machine_101 = MachineScheduleResult(
        order_line_id=101,
        contract_no="HT101",
        order_no="SO101",
        product_model="MC1-80",
        schedule_status="scheduled",
    )
    machine_102 = MachineScheduleResult(
        order_line_id=102,
        contract_no="HT102",
        order_no="SO102",
        product_model="MC2-100",
        schedule_status="scheduled",
    )
    db_session.add_all([machine_101, machine_102])
    await db_session.flush()

    db_session.add_all(
        [
            PartScheduleResult(
                order_line_id=101,
                machine_schedule_id=machine_101.id,
                assembly_name="机身",
                parent_material_no="ASM_101",
                parent_name="机身",
                node_level=1,
                bom_path="机身(ASM_101) / Part 101(P101)",
                bom_path_key="root:ASM_101>1",
                production_sequence=2,
                part_material_no="P101",
                part_name="Part 101",
                is_key_part=True,
            ),
            PartScheduleResult(
                order_line_id=101,
                machine_schedule_id=machine_101.id,
                assembly_name="机身",
                parent_material_no="ASM_101",
                parent_name="机身",
                node_level=1,
                bom_path="机身(ASM_101) / Part 100(P100)",
                bom_path_key="root:ASM_101>2",
                production_sequence=1,
                part_material_no="P100",
                part_name="Part 100",
                is_key_part=False,
            ),
            PartScheduleResult(
                order_line_id=102,
                machine_schedule_id=machine_102.id,
                assembly_name="机身",
                production_sequence=1,
                part_material_no="P102",
                part_name="Part 102",
                is_key_part=True,
            ),
        ]
    )
    await db_session.commit()

    resp = await app_client.get(
        "/api/exports/part-schedules?order_line_id=101&sort_field=production_sequence&sort_order=asc"
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert len(resp.content) > 0
    assert "filename*=UTF-8''" in resp.headers["content-disposition"]
    assert "%E9%9B%B6%E4%BB%B6%E6%8E%92%E4%BA%A7%E6%98%8E%E7%BB%86_" in resp.headers["content-disposition"]

    workbook = load_workbook(BytesIO(resp.content))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    first_row = [cell.value for cell in sheet[2]]
    second_row = [cell.value for cell in sheet[3]]

    assert headers[:7] == ["订单行ID", "合同编号", "产品型号", "工厂", "订单编号", "部装名称", "上级物料号"]
    assert "部装装配时长(天)" in headers
    assert "零件自身周期(天)" in headers
    assert "关键件倒排周期(天)" in headers
    assert "完整路径" in headers
    assert "关键件标识" in headers
    assert first_row[0] == 101
    assert first_row[10] == 1
    assert second_row[10] == 2
    assert first_row[9] == "机身(ASM_101) / Part 100(P100)"
    assert "否" in first_row
    assert "是" in second_row


@pytest.mark.asyncio
async def test_export_part_schedules_batches_parts_for_multiple_snapshots(app_client, db_session):
    db_session.add_all(
        [
            OrderScheduleSnapshot(
                order_line_id=201,
                contract_no="HT201",
                order_no="SO201",
                product_model="MC2-100",
                plant="1100",
                schedule_status="scheduled",
                warning_level="normal",
            ),
            OrderScheduleSnapshot(
                order_line_id=202,
                contract_no="HT202",
                order_no="SO202",
                product_model="MC1-80",
                plant="1200",
                schedule_status="scheduled",
                warning_level="normal",
            ),
            PartScheduleResult(
                order_line_id=201,
                assembly_name="Assembly-201",
                bom_path="Assembly-201(ASM201) / Part 201-B(P201-B)",
                production_sequence=2,
                part_material_no="P201-B",
                part_name="Part 201-B",
                is_key_part=False,
            ),
            PartScheduleResult(
                order_line_id=201,
                assembly_name="Assembly-201",
                bom_path="Assembly-201(ASM201) / Part 201-A(P201-A)",
                production_sequence=1,
                part_material_no="P201-A",
                part_name="Part 201-A",
                is_key_part=True,
            ),
            PartScheduleResult(
                order_line_id=202,
                assembly_name="Assembly-202",
                production_sequence=1,
                part_material_no="P202-A",
                part_name="Part 202-A",
                is_key_part=True,
            ),
        ]
    )
    await db_session.commit()

    resp = await app_client.get("/api/exports/part-schedules?sort_field=contract_no&sort_order=asc")
    assert resp.status_code == 200

    workbook = load_workbook(BytesIO(resp.content))
    sheet = workbook.active
    row_2 = [cell.value for cell in sheet[2]]
    row_3 = [cell.value for cell in sheet[3]]
    row_4 = [cell.value for cell in sheet[4]]

    assert row_2[0] == 201
    assert row_2[1] == "HT201"
    assert row_2[3] == "1100"
    assert row_2[4] == "SO201"
    assert row_2[10] == 1
    assert row_3[0] == 201
    assert row_3[10] == 2
    assert row_4[0] == 202
    assert row_4[1] == "HT202"


@pytest.mark.asyncio
async def test_export_part_schedules_hides_placeholder_rows(app_client, db_session):
    db_session.add_all(
        [
            OrderScheduleSnapshot(
                order_line_id=301,
                contract_no="HT301",
                order_no="SO301",
                product_model="MC1-80",
                plant="1100",
                schedule_status="scheduled",
                warning_level="normal",
            ),
            PartScheduleResult(
                order_line_id=301,
                assembly_name="平衡缸",
                production_sequence=1,
                planned_end_date=datetime(2025, 3, 20, 0, 0, 0),
            ),
            PartScheduleResult(
                order_line_id=301,
                assembly_name="机身",
                production_sequence=2,
                part_material_no="P301",
                part_name="Part 301",
                is_key_part=True,
            ),
        ]
    )
    await db_session.commit()

    resp = await app_client.get("/api/exports/part-schedules?order_line_id=301")
    assert resp.status_code == 200

    workbook = load_workbook(BytesIO(resp.content))
    sheet = workbook.active

    assert sheet.max_row == 2
    assert sheet[2][5].value == "机身"
    assert sheet[2][13].value == "Part 301"


@pytest.mark.asyncio
async def test_export_part_schedules_filters_by_snapshot_contract_without_machine_rows(app_client, db_session):
    db_session.add_all(
        [
            OrderScheduleSnapshot(
                order_line_id=401,
                contract_no="HT401",
                order_no="SO401",
                product_model="MC4-401",
                plant="1100",
                schedule_status="scheduled",
                warning_level="normal",
            ),
            OrderScheduleSnapshot(
                order_line_id=402,
                contract_no="HT402",
                order_no="SO402",
                product_model="MC4-402",
                plant="1200",
                schedule_status="scheduled",
                warning_level="normal",
            ),
            PartScheduleResult(
                order_line_id=401,
                assembly_name="Assembly-401",
                production_sequence=1,
                part_material_no="P401",
                part_name="Part 401",
                is_key_part=True,
            ),
            PartScheduleResult(
                order_line_id=402,
                assembly_name="Assembly-402",
                production_sequence=1,
                part_material_no="P402",
                part_name="Part 402",
                is_key_part=True,
            ),
        ]
    )
    await db_session.commit()

    resp = await app_client.get("/api/exports/part-schedules?contract_no=HT401")
    assert resp.status_code == 200

    workbook = load_workbook(BytesIO(resp.content))
    sheet = workbook.active

    assert sheet.max_row == 2
    assert sheet[2][0].value == 401
    assert sheet[2][1].value == "HT401"
    assert sheet[2][4].value == "SO401"


@pytest.mark.asyncio
async def test_export_part_schedules_filters_by_part_warning_level_and_planned_end_date(app_client, db_session):
    db_session.add_all(
        [
            OrderScheduleSnapshot(
                order_line_id=470,
                contract_no="HT470",
                order_no="SO470",
                product_model="MC4-470",
                plant="1100",
                schedule_status="scheduled",
                warning_level="normal",
                confirmed_delivery_date=datetime(2025, 5, 20, 0, 0, 0),
            ),
            OrderScheduleSnapshot(
                order_line_id=471,
                contract_no="HT471",
                order_no="SO471",
                product_model="MC4-471",
                plant="1100",
                schedule_status="scheduled",
                warning_level="abnormal",
                confirmed_delivery_date=datetime(2025, 3, 5, 0, 0, 0),
            ),
            PartScheduleResult(
                order_line_id=470,
                assembly_name="Assembly-470",
                production_sequence=1,
                part_material_no="P470",
                part_name="Part 470",
                is_key_part=True,
                warning_level="abnormal",
                planned_end_date=datetime(2025, 3, 18, 0, 0, 0),
            ),
            PartScheduleResult(
                order_line_id=471,
                assembly_name="Assembly-471",
                production_sequence=1,
                part_material_no="P471",
                part_name="Part 471",
                is_key_part=True,
                warning_level="normal",
                planned_end_date=datetime(2025, 4, 18, 0, 0, 0),
            ),
        ]
    )
    await db_session.commit()

    resp = await app_client.get(
        "/api/exports/part-schedules?warning_level=abnormal&date_from=2025-03-01&date_to=2025-03-31"
    )
    assert resp.status_code == 200

    workbook = load_workbook(BytesIO(resp.content))
    sheet = workbook.active

    assert sheet.max_row == 2
    assert sheet[2][0].value == 470
    assert sheet[2][12].value == "P470"


@pytest.mark.asyncio
async def test_export_part_schedules_service_filters_by_part_fields(db_session, monkeypatch):
    db_session.add_all(
        [
            OrderScheduleSnapshot(
                order_line_id=451,
                contract_no="HT451",
                order_no="SO451",
                product_model="MC4-451",
                plant="1100",
                schedule_status="scheduled",
                warning_level="normal",
            ),
            OrderScheduleSnapshot(
                order_line_id=452,
                contract_no="HT452",
                order_no="SO452",
                product_model="MC4-452",
                plant="1200",
                schedule_status="scheduled",
                warning_level="normal",
            ),
            PartScheduleResult(
                order_line_id=451,
                assembly_name="电控",
                production_sequence=1,
                part_material_no="P451",
                part_name="Part 451",
                is_key_part=True,
            ),
            PartScheduleResult(
                order_line_id=452,
                assembly_name="机身",
                production_sequence=1,
                part_material_no="P452",
                part_name="Part 452",
                is_key_part=True,
            ),
        ]
    )
    await db_session.commit()

    service = ExportService(db_session)

    async def fake_ensure_ready():
        return None

    monkeypatch.setattr(service, "_ensure_ready", fake_ensure_ready)

    buf, _, _ = await service.export_part_schedules(assembly_name="电控")
    workbook = load_workbook(buf)
    sheet = workbook.active

    assert sheet.max_row == 2
    assert sheet[2][0].value == 451
    assert sheet[2][5].value == "电控"
    assert sheet[2][13].value == "Part 451"


def _build_fake_snapshot(order_line_id: int) -> SimpleNamespace:
    return SimpleNamespace(
        order_line_id=order_line_id,
        contract_no=f"HT{order_line_id:05d}",
        customer_name=f"Customer {order_line_id}",
        product_series="MC",
        product_model="MC-100",
        product_name="Machine",
        material_no=f"MAT-{order_line_id:05d}",
        plant="1000",
        quantity=Decimal("1"),
        order_type="1",
        order_date=datetime(2026, 3, 1, 0, 0, 0),
        line_total_amount=Decimal("100"),
        business_group="BG",
        custom_no=f"DZ-{order_line_id:05d}",
        sales_person_name="Tester",
        sales_branch_company="Branch",
        sales_sub_branch="SubBranch",
        order_no=f"SO{order_line_id:05d}",
        sap_code=f"SAP{order_line_id:05d}",
        sap_line_no="000010",
        confirmed_delivery_date=datetime(2026, 4, 1, 0, 0, 0),
        drawing_released=True,
        drawing_release_date=datetime(2026, 3, 2, 0, 0, 0),
        custom_requirement=None,
        review_comment=None,
        trigger_date=datetime(2026, 3, 3, 0, 0, 0),
        schedule_status="scheduled",
        planned_start_date=datetime(2026, 3, 4, 0, 0, 0),
        planned_end_date=datetime(2026, 4, 1, 0, 0, 0),
        machine_cycle_days=Decimal("10"),
        machine_assembly_days=Decimal("2"),
        warning_level="normal",
    )


def _build_fake_part(order_line_id: int) -> SimpleNamespace:
    return SimpleNamespace(
        order_line_id=order_line_id,
        assembly_name="Assembly",
        parent_material_no=None,
        parent_name=None,
        node_level=None,
        bom_path=None,
        production_sequence=1,
        assembly_time_days=Decimal("2"),
        part_material_no=f"P-{order_line_id:05d}",
        part_name=f"Part {order_line_id}",
        is_key_part=bool(order_line_id % 2),
        part_cycle_days=Decimal("5"),
        key_part_material_no=None,
        key_part_name=None,
        key_part_raw_material_desc=None,
        key_part_cycle_days=None,
        planned_start_date=datetime(2026, 3, 4, 0, 0, 0),
        planned_end_date=datetime(2026, 3, 9, 0, 0, 0),
    )


@pytest.mark.asyncio
async def test_export_machine_schedules_batches_snapshot_export(db_session, monkeypatch):
    from app.config import settings

    service = ExportService(db_session)
    snapshots = [_build_fake_snapshot(index) for index in range(1, 10002)]

    async def fake_ensure_ready():
        return None

    async def fail_paginate(*args, **kwargs):
        raise AssertionError("paginate should not be used for machine export")

    async def fake_stream_for_export_batches(*, batch_size, **kwargs):
        for start in range(0, len(snapshots), batch_size):
            yield snapshots[start : start + batch_size]

    async def fail_list_for_export_batch(*args, **kwargs):
        raise AssertionError("offset pagination should not be used for machine export")

    monkeypatch.setattr(service, "_ensure_ready", fake_ensure_ready)
    monkeypatch.setattr(service.snapshot_repo, "paginate", fail_paginate)
    monkeypatch.setattr(service.snapshot_repo, "stream_for_export_batches", fake_stream_for_export_batches)
    monkeypatch.setattr(service.snapshot_repo, "list_for_export_batch", fail_list_for_export_batch)

    original_limit = settings.export_excel_max_rows
    settings.export_excel_max_rows = 20000
    try:
        buf, _, _ = await service.export_machine_schedules()
        workbook = load_workbook(buf)
        sheet = workbook.active

        assert sheet.max_row == 10002
        assert sheet.cell(row=10002, column=1).value == 10001
    finally:
        settings.export_excel_max_rows = original_limit


@pytest.mark.asyncio
async def test_export_part_schedules_skips_snapshot_sort_when_part_sort_requested(db_session, monkeypatch):
    service = ExportService(db_session)
    captured_kwargs = {}

    async def fake_ensure_ready():
        return None

    async def fake_stream_for_export_rows(*, batch_size, **kwargs):
        captured_kwargs["batch_size"] = batch_size
        captured_kwargs.update(kwargs)
        yield [(_build_fake_snapshot(1), _build_fake_part(1))]

    async def fail_stream_for_export_batches(*args, **kwargs):
        raise AssertionError("snapshot batch export should not be used for part export")

    async def fail_find_by_order_line_ids(*args, **kwargs):
        raise AssertionError("find_by_order_line_ids should not be used for part export")

    monkeypatch.setattr(service, "_ensure_ready", fake_ensure_ready)
    monkeypatch.setattr(service.snapshot_repo, "stream_for_export_batches", fail_stream_for_export_batches)
    monkeypatch.setattr(service.psr_repo, "find_by_order_line_ids", fail_find_by_order_line_ids)
    monkeypatch.setattr(service.psr_repo, "stream_for_export_rows", fake_stream_for_export_rows)

    await service.export_part_schedules(sort_field="production_sequence", sort_order="asc")

    assert captured_kwargs["snapshot_sort_field"] is None
    assert captured_kwargs["snapshot_sort_order"] is None
    assert captured_kwargs["part_sort_field"] == "production_sequence"
    assert captured_kwargs["part_sort_order"] == "asc"


@pytest.mark.asyncio
async def test_export_part_schedules_batches_snapshot_export(db_session, monkeypatch):
    from app.config import settings

    service = ExportService(db_session)
    export_rows = [(_build_fake_snapshot(index), _build_fake_part(index)) for index in range(1, 10002)]

    async def fake_ensure_ready():
        return None

    async def fail_paginate(*args, **kwargs):
        raise AssertionError("paginate should not be used for part export")

    async def fail_stream_for_export_batches(*args, **kwargs):
        raise AssertionError("snapshot batch export should not be used for part export")

    async def fail_find_by_order_line_ids(*args, **kwargs):
        raise AssertionError("find_by_order_line_ids should not be used for part export")

    async def fake_stream_for_export_rows(*, batch_size, **kwargs):
        for start in range(0, len(export_rows), batch_size):
            batch = export_rows[start : start + batch_size]
            assert len(batch) <= settings.export_batch_size
            yield batch

    monkeypatch.setattr(service, "_ensure_ready", fake_ensure_ready)
    monkeypatch.setattr(service.snapshot_repo, "paginate", fail_paginate)
    monkeypatch.setattr(service.snapshot_repo, "stream_for_export_batches", fail_stream_for_export_batches)
    monkeypatch.setattr(service.psr_repo, "find_by_order_line_ids", fail_find_by_order_line_ids)
    monkeypatch.setattr(service.psr_repo, "stream_for_export_rows", fake_stream_for_export_rows)

    original_limit = settings.export_excel_max_rows
    settings.export_excel_max_rows = 20000
    try:
        buf, _, _ = await service.export_part_schedules()
        workbook = load_workbook(buf)
        sheet = workbook.active

        assert sheet.max_row == 10002
        assert sheet.cell(row=10002, column=1).value == 10001
    finally:
        settings.export_excel_max_rows = original_limit


@pytest.mark.asyncio
async def test_export_machine_schedules_supports_csv(app_client, db_session):
    db_session.add(
        OrderScheduleSnapshot(
            order_line_id=901,
            contract_no="HT901",
            order_no="SO901",
            product_model="MC-901",
            schedule_status="scheduled",
            warning_level="normal",
        )
    )
    await db_session.commit()

    resp = await app_client.get("/api/exports/machine-schedules?export_format=csv")

    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]
    assert "filename*=UTF-8''" in resp.headers["content-disposition"]
    assert "订单行ID" in resp.text
    assert "HT901" in resp.text


@pytest.mark.asyncio
async def test_export_machine_schedules_rejects_large_xlsx_export(app_client, db_session, monkeypatch):
    from app.config import settings

    original_limit = settings.export_excel_max_rows
    settings.export_excel_max_rows = 1
    try:
        db_session.add_all(
            [
                OrderScheduleSnapshot(order_line_id=1001, contract_no="HT1001", schedule_status="scheduled"),
                OrderScheduleSnapshot(order_line_id=1002, contract_no="HT1002", schedule_status="scheduled"),
            ]
        )
        await db_session.commit()

        resp = await app_client.get("/api/exports/machine-schedules")

        assert resp.status_code == 200
        body = resp.json()
        assert body["code"] == 4003
        assert "请改用 CSV 导出" in body["message"]
    finally:
        settings.export_excel_max_rows = original_limit


@pytest.mark.asyncio
async def test_machine_xlsx_limit_check_uses_limit_probe_instead_of_full_count(db_session, monkeypatch):
    service = ExportService(db_session)
    captured: dict[str, object] = {}

    async def fake_has_export_rows_beyond_limit(*, max_rows: int, **kwargs):
        captured["max_rows"] = max_rows
        captured["filters"] = kwargs
        return False

    async def fail_list_for_export_batch(**kwargs):
        raise AssertionError("xlsx row limit check should not scan export batches")

    async def fail_count_for_export(**kwargs):
        raise AssertionError("xlsx row limit check should not use full count query")

    monkeypatch.setattr(service.snapshot_repo, "has_export_rows_beyond_limit", fake_has_export_rows_beyond_limit)
    monkeypatch.setattr(service.snapshot_repo, "count_for_export", fail_count_for_export)
    monkeypatch.setattr(service.snapshot_repo, "list_for_export_batch", fail_list_for_export_batch)

    await service._ensure_machine_xlsx_within_limit({})

    assert captured["max_rows"] >= 1
    assert captured["filters"] == {}


@pytest.mark.asyncio
async def test_part_xlsx_limit_check_uses_limit_probe_instead_of_full_count(db_session, monkeypatch):
    service = ExportService(db_session)
    captured: dict[str, object] = {}

    async def fake_has_export_rows_beyond_limit(*, max_rows: int, **kwargs):
        captured["max_rows"] = max_rows
        captured["filters"] = kwargs
        return False

    def fail_iter_part_export_rows(*args, **kwargs):
        raise AssertionError("xlsx row limit check should not iterate export rows")

    async def fail_count_for_export(**kwargs):
        raise AssertionError("xlsx row limit check should not use full count query")

    monkeypatch.setattr(service.psr_repo, "has_export_rows_beyond_limit", fake_has_export_rows_beyond_limit)
    monkeypatch.setattr(service.psr_repo, "count_for_export", fail_count_for_export)
    monkeypatch.setattr(service, "_iter_part_export_rows_async", fail_iter_part_export_rows)

    await service._ensure_part_xlsx_within_limit({})

    assert captured["max_rows"] >= 1
    assert captured["filters"] == {}

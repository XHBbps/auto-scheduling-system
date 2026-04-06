import csv
import io
import logging
import tempfile
from collections.abc import Iterable
from inspect import isawaitable
from time import perf_counter
from typing import Any, BinaryIO

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.common.datetime_utils import utc_now
from app.common.enums import ScheduleStatus, WarningLevel
from app.common.exceptions import BizException, ErrorCode
from app.config import settings
from app.repository.order_schedule_snapshot_repo import OrderScheduleSnapshotRepo
from app.repository.part_schedule_result_repo import PartScheduleResultRepo
from app.services.schedule_snapshot_refresh_service import ScheduleSnapshotRefreshService

logger = logging.getLogger(__name__)

SCHEDULE_STATUS_LABELS = {
    ScheduleStatus.SCHEDULED: "已排产",
    ScheduleStatus.SCHEDULED_STALE: "待重排",
    ScheduleStatus.PENDING_DELIVERY: "待交期",
    ScheduleStatus.PENDING_DRAWING: "待发图",
    ScheduleStatus.PENDING_TRIGGER: "待触发",
    ScheduleStatus.SCHEDULABLE: "可排产",
    ScheduleStatus.MISSING_BOM: "缺少BOM",
}

WARNING_LEVEL_LABELS = {
    WarningLevel.NORMAL: "正常",
    WarningLevel.ABNORMAL: "异常",
}

ORDER_TYPE_LABELS = {
    "1": "常规",
    "2": "选配",
    "3": "定制",
}

MACHINE_SCHEDULE_HEADERS = [
    "订单行ID",
    "合同编号",
    "客户名称",
    "产品系列",
    "产品型号",
    "产品名称",
    "整机物料号",
    "工厂",
    "数量",
    "订单类型",
    "订单日期",
    "合同金额",
    "事业群",
    "定制号",
    "销售人员",
    "分公司",
    "支公司",
    "订单编号",
    "SAP编码",
    "SAP行号",
    "确认交货期",
    "发图状态",
    "发图日期",
    "定制要求",
    "评审意见",
    "触发日期",
    "排产状态",
    "计划开工日",
    "计划完工日",
    "整机主周期(天)",
    "整机总装时长(天)",
    "异常标识",
]

PART_SCHEDULE_HEADERS = [
    "订单行ID",
    "合同编号",
    "产品型号",
    "工厂",
    "订单编号",
    "部装名称",
    "上级物料号",
    "上级名称",
    "层级",
    "完整路径",
    "生产顺序",
    "部装装配时长(天)",
    "零件物料号",
    "零件名称",
    "关键件标识",
    "零件自身周期(天)",
    "关键件物料号",
    "关键件名称",
    "关键件毛坯描述",
    "关键件倒排周期(天)",
    "计划开工日",
    "计划完工日",
]


class ExportService:
    def __init__(self, session: AsyncSession):
        self.session = session
        self.snapshot_repo = OrderScheduleSnapshotRepo(session)
        self.psr_repo = PartScheduleResultRepo(session)
        self.snapshot_refresh_service = ScheduleSnapshotRefreshService(session)

    @staticmethod
    async def _resolve_result(result: Any) -> Any:
        if isawaitable(result):
            return await result
        return result

    async def _ensure_ready(self) -> None:
        await self.snapshot_refresh_service.ensure_seeded(
            source="export_service",
            reason="lazy_snapshot_seed",
        )

    @staticmethod
    def _build_snapshot_export_filters(filters: dict[str, Any]) -> dict[str, Any]:
        return {
            key: value
            for key, value in filters.items()
            if key
            in {
                "order_line_id",
                "contract_no",
                "customer_name",
                "product_series",
                "product_model",
                "plant",
                "order_no",
                "schedule_status",
                "warning_level",
                "date_from",
                "date_to",
                "sort_field",
                "sort_order",
            }
        }

    @staticmethod
    def _build_part_export_sort(sort_field: str | None, sort_order: str | None) -> tuple[str | None, str | None]:
        if sort_field not in PartScheduleResultRepo.PART_FIELDS:
            return None, None
        return sort_field, sort_order

    async def export_machine_schedules(
        self,
        *,
        export_format: str = "xlsx",
        **filters: Any,
    ) -> tuple[BinaryIO, str, str]:
        started_perf = perf_counter()
        await self._ensure_ready()
        filters.pop("page_no", None)
        filters.pop("page_size", None)
        buffer = self._build_spooled_buffer()
        filter_summary = self._summarize_filters(filters)
        logger.info("Machine schedule export started: format=%s filters=%s", export_format, filter_summary)
        try:
            if export_format == "csv":
                row_count = await self._resolve_result(self._write_machine_csv(buffer, filters))
                filename = f"整机排产列表_{utc_now().strftime('%Y%m%d%H%M%S')}.csv"
                logger.info(
                    "Machine schedule export finished: format=%s row_count=%s duration_ms=%s filters=%s",
                    export_format,
                    row_count,
                    round((perf_counter() - started_perf) * 1000, 2),
                    filter_summary,
                )
                return buffer, filename, "text/csv; charset=utf-8"

            await self._resolve_result(self._ensure_machine_xlsx_within_limit(filters))
            row_count = await self._resolve_result(self._write_machine_xlsx(buffer, filters))
            filename = f"整机排产列表_{utc_now().strftime('%Y%m%d%H%M%S')}.xlsx"
            logger.info(
                "Machine schedule export finished: format=%s row_count=%s duration_ms=%s filters=%s",
                export_format,
                row_count,
                round((perf_counter() - started_perf) * 1000, 2),
                filter_summary,
            )
            return buffer, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        except Exception:
            logger.exception(
                "Machine schedule export failed: format=%s duration_ms=%s filters=%s",
                export_format,
                round((perf_counter() - started_perf) * 1000, 2),
                filter_summary,
            )
            raise

    async def export_part_schedules(
        self,
        *,
        export_format: str = "xlsx",
        **filters: Any,
    ) -> tuple[BinaryIO, str, str]:
        started_perf = perf_counter()
        await self._ensure_ready()
        filters.pop("page_no", None)
        filters.pop("page_size", None)
        buffer = self._build_spooled_buffer()
        filter_summary = self._summarize_filters(filters)
        logger.info("Part schedule export started: format=%s filters=%s", export_format, filter_summary)
        try:
            if export_format == "csv":
                row_count = await self._resolve_result(self._write_part_csv(buffer, filters))
                filename = f"零件排产明细_{utc_now().strftime('%Y%m%d%H%M%S')}.csv"
                logger.info(
                    "Part schedule export finished: format=%s row_count=%s duration_ms=%s filters=%s",
                    export_format,
                    row_count,
                    round((perf_counter() - started_perf) * 1000, 2),
                    filter_summary,
                )
                return buffer, filename, "text/csv; charset=utf-8"

            await self._resolve_result(self._ensure_part_xlsx_within_limit(filters))
            row_count = await self._resolve_result(self._write_part_xlsx(buffer, filters))
            filename = f"零件排产明细_{utc_now().strftime('%Y%m%d%H%M%S')}.xlsx"
            logger.info(
                "Part schedule export finished: format=%s row_count=%s duration_ms=%s filters=%s",
                export_format,
                row_count,
                round((perf_counter() - started_perf) * 1000, 2),
                filter_summary,
            )
            return buffer, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        except Exception:
            logger.exception(
                "Part schedule export failed: format=%s duration_ms=%s filters=%s",
                export_format,
                round((perf_counter() - started_perf) * 1000, 2),
                filter_summary,
            )
            raise

    def _validate_excel_row_limit(self, row_count: int, *, export_format: str) -> None:
        if export_format != "xlsx":
            return
        if row_count <= settings.export_excel_max_rows:
            return
        raise BizException(
            ErrorCode.BIZ_VALIDATION_FAILED,
            f"Excel 导出最多支持 {settings.export_excel_max_rows} 行，请改用 CSV 导出。",
        )

    @staticmethod
    def _build_spooled_buffer() -> BinaryIO:
        return tempfile.SpooledTemporaryFile(
            max_size=settings.export_spool_max_size_bytes,
            mode="w+b",
        )

    @staticmethod
    def _summarize_filters(filters: dict[str, Any]) -> dict[str, Any]:
        return {key: value for key, value in filters.items() if value not in (None, "", [], {}, ())}

    @staticmethod
    def _build_xlsx_buffer(sheet_name: str, headers: list[str], rows: Iterable[list[Any]]) -> BinaryIO:
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)

        if "Sheet" in workbook.sheetnames:
            default_sheet = workbook["Sheet"]
            workbook.remove(default_sheet)

        buffer = ExportService._build_spooled_buffer()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _build_csv_buffer(headers: list[str], rows: Iterable[list[Any]]) -> BinaryIO:
        byte_buffer = ExportService._build_spooled_buffer()
        text_buffer = io.TextIOWrapper(byte_buffer, encoding="utf-8-sig", newline="", write_through=True)
        writer = csv.writer(text_buffer)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        text_buffer.flush()
        text_buffer.detach()
        byte_buffer.seek(0)
        return byte_buffer

    async def _write_machine_xlsx(self, buffer: BinaryIO, filters: dict[str, Any]) -> int:
        return await self._write_machine_xlsx_async(buffer, filters)

    async def _write_machine_xlsx_async(self, buffer: BinaryIO, filters: dict[str, Any]) -> int:
        workbook = Workbook(write_only=True)
        row_count = 0
        try:
            worksheet = workbook.create_sheet(title="machine_schedule")
            worksheet.append(MACHINE_SCHEDULE_HEADERS)
            async for batch in self._iter_snapshot_batches_async(filters):
                for item in batch:
                    worksheet.append(self._serialize_machine_schedule_row(item))
                    row_count += 1
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])
            workbook.save(buffer)
            buffer.seek(0)
            return row_count
        finally:
            workbook.close()

    async def _write_machine_csv(self, buffer: BinaryIO, filters: dict[str, Any]) -> int:
        return await self._write_machine_csv_async(buffer, filters)

    async def _write_machine_csv_async(self, buffer: BinaryIO, filters: dict[str, Any]) -> int:
        text_buffer = io.TextIOWrapper(buffer, encoding="utf-8-sig", newline="", write_through=True)
        writer = csv.writer(text_buffer)
        row_count = 0
        writer.writerow(MACHINE_SCHEDULE_HEADERS)
        async for batch in self._iter_snapshot_batches_async(filters):
            for item in batch:
                writer.writerow(self._serialize_machine_schedule_row(item))
                row_count += 1
        text_buffer.flush()
        text_buffer.detach()
        buffer.seek(0)
        return row_count

    async def _write_part_xlsx(self, buffer: BinaryIO, filters: dict[str, Any]) -> int:
        return await self._write_part_xlsx_async(buffer, filters)

    async def _write_part_xlsx_async(self, buffer: BinaryIO, filters: dict[str, Any]) -> int:
        workbook = Workbook(write_only=True)
        row_count = 0
        try:
            worksheet = workbook.create_sheet(title="part_schedule")
            worksheet.append(PART_SCHEDULE_HEADERS)
            async for snapshot, part in self._iter_part_export_rows_async(filters):
                worksheet.append(self._serialize_part_schedule_row(snapshot, part))
                row_count += 1
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])
            workbook.save(buffer)
            buffer.seek(0)
            return row_count
        finally:
            workbook.close()

    async def _ensure_machine_xlsx_within_limit(self, filters: dict[str, Any]) -> None:
        exceeds_limit = await self.snapshot_repo.has_export_rows_beyond_limit(
            max_rows=settings.export_excel_max_rows,
            **self._build_snapshot_export_filters(dict(filters)),
        )
        if exceeds_limit:
            self._validate_excel_row_limit(settings.export_excel_max_rows + 1, export_format="xlsx")

    async def _ensure_part_xlsx_within_limit(self, filters: dict[str, Any]) -> None:
        exceeds_limit = await self.psr_repo.has_export_rows_beyond_limit(
            max_rows=settings.export_excel_max_rows,
            **dict(filters),
        )
        if exceeds_limit:
            self._validate_excel_row_limit(settings.export_excel_max_rows + 1, export_format="xlsx")

    async def _write_part_csv(self, buffer: BinaryIO, filters: dict[str, Any]) -> int:
        return await self._write_part_csv_async(buffer, filters)

    async def _write_part_csv_async(self, buffer: BinaryIO, filters: dict[str, Any]) -> int:
        text_buffer = io.TextIOWrapper(buffer, encoding="utf-8-sig", newline="", write_through=True)
        writer = csv.writer(text_buffer)
        row_count = 0
        writer.writerow(PART_SCHEDULE_HEADERS)
        async for snapshot, part in self._iter_part_export_rows_async(filters):
            writer.writerow(self._serialize_part_schedule_row(snapshot, part))
            row_count += 1
        text_buffer.flush()
        text_buffer.detach()
        buffer.seek(0)
        return row_count

    async def _iter_snapshot_batches_async(self, filters: dict[str, Any]):
        async for batch in self.snapshot_repo.stream_for_export_batches(
            batch_size=settings.export_batch_size,
            **dict(filters),
        ):
            yield batch

    async def _iter_part_export_rows_async(self, filters: dict[str, Any]):
        sort_field = filters.get("sort_field")
        sort_order = filters.get("sort_order")
        snapshot_sort_field = sort_field if sort_field in PartScheduleResultRepo.SNAPSHOT_FIELDS else None
        part_sort_field, part_sort_order = self._build_part_export_sort(sort_field, sort_order)
        snapshot_sort_order = sort_order if snapshot_sort_field else None

        async for batch in self.psr_repo.stream_for_export_rows(
            batch_size=settings.export_batch_size,
            snapshot_sort_field=snapshot_sort_field,
            snapshot_sort_order=snapshot_sort_order,
            part_sort_field=part_sort_field,
            part_sort_order=part_sort_order,
            **dict(filters),
        ):
            for snapshot, part in batch:
                yield snapshot, part

    def _serialize_machine_schedule_row(self, item: Any) -> list[Any]:
        return [
            item.order_line_id,
            item.contract_no,
            item.customer_name,
            item.product_series,
            item.product_model,
            item.product_name,
            item.material_no,
            item.plant,
            float(item.quantity) if item.quantity is not None else None,
            self._format_order_type(item.order_type),
            item.order_date.strftime("%Y-%m-%d") if item.order_date else None,
            float(item.line_total_amount) if item.line_total_amount is not None else None,
            item.business_group,
            item.custom_no,
            item.sales_person_name,
            item.sales_branch_company,
            item.sales_sub_branch,
            item.order_no,
            item.sap_code,
            item.sap_line_no,
            item.confirmed_delivery_date.strftime("%Y-%m-%d") if item.confirmed_delivery_date else None,
            self._format_drawing_released(item.drawing_released),
            item.drawing_release_date.strftime("%Y-%m-%d") if item.drawing_release_date else None,
            item.custom_requirement,
            item.review_comment,
            item.trigger_date.strftime("%Y-%m-%d") if item.trigger_date else None,
            self._format_schedule_status(item.schedule_status),
            item.planned_start_date.strftime("%Y-%m-%d") if item.planned_start_date else None,
            item.planned_end_date.strftime("%Y-%m-%d") if item.planned_end_date else None,
            float(item.machine_cycle_days) if item.machine_cycle_days is not None else None,
            float(item.machine_assembly_days) if item.machine_assembly_days is not None else None,
            self._format_warning_level(item.warning_level),
        ]

    def _serialize_part_schedule_row(self, snapshot: Any, part: Any) -> list[Any]:
        return [
            snapshot.order_line_id,
            snapshot.contract_no,
            snapshot.product_model,
            snapshot.plant,
            snapshot.order_no,
            part.assembly_name,
            part.parent_material_no,
            part.parent_name,
            part.node_level,
            part.bom_path,
            part.production_sequence,
            float(part.assembly_time_days) if part.assembly_time_days is not None else None,
            part.part_material_no,
            part.part_name,
            "是" if part.is_key_part else "否",
            float(part.part_cycle_days) if part.part_cycle_days is not None else None,
            part.key_part_material_no,
            part.key_part_name,
            part.key_part_raw_material_desc,
            float(part.key_part_cycle_days) if part.key_part_cycle_days is not None else None,
            part.planned_start_date.strftime("%Y-%m-%d") if part.planned_start_date else None,
            part.planned_end_date.strftime("%Y-%m-%d") if part.planned_end_date else None,
        ]

    @staticmethod
    def _format_schedule_status(value: str | None) -> str | None:
        if not value:
            return value
        return SCHEDULE_STATUS_LABELS.get(value, value)

    @staticmethod
    def _format_warning_level(value: str | None) -> str | None:
        if not value:
            return value
        return WARNING_LEVEL_LABELS.get(value, value)

    @staticmethod
    def _format_drawing_released(value: bool | None) -> str | None:
        if value is None:
            return None
        return "已发图" if value else "未发图"

    @staticmethod
    def _format_order_type(value: str | None) -> str | None:
        if not value:
            return value
        return ORDER_TYPE_LABELS.get(value, value)
